from __future__ import print_function
import pickle
import tkinter
import json
from slugify import slugify
import math
from collections import OrderedDict
import os.path
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import pprint
from datetime import datetime, timezone
import iso8601
import numpy
import matplotlib.pyplot as plt
import requests
import os.path
import pandas as pd
from pathlib import Path

pp = pprint.PrettyPrinter(indent=4);
BINS=400
# If modifying these scopes, delete the file token.pickle.
SCOPES = ['https://www.googleapis.com/auth/documents.readonly', 'https://www.googleapis.com/auth/drive.readonly', 'https://www.googleapis.com/auth/drive.activity.readonly']
ITEMPERREQ = 1000

# The ID of a sample document.
DOCUMENT_ID = '0B4Fujvv5MfqbNGs3NjV6RndYOVk'
DIR = 'data/'

def listFiles (id, drive, folders, files, nextToken = None, callingName = "", filesParents = None):
    print("list %s"%id)
    q = "'" + id + "' in parents";
    if nextToken == None:
        response = drive.files().list(q=q, pageSize = 1000).execute()
    else:
        response = drive.files().list(q=q, pageSize = 1000, pageToken = nextToken).execute()
    for file in response.get('files', []):
        # Process change
        if(file.get("mimeType") == "application/vnd.google-apps.folder"):

            folders.append((file.get("id"),callingName+"/"+file.get("name")))
           # folders[file.get("id")] = files[file.get("name")] # Reference to scope of map to add
          
        elif (file.get("mimeType") == "application/vnd.google-apps.document"):
            files[file.get("id")] = file;
            if(filesParents!=None):
                filesParents[file.get("id")] = callingName + "/"+file.get("name")

    print("token %s"%response.get('nextPageToken'))
    if(response.get('nextPageToken')!=("" or None)):
        print("next token %s"%response.get('nextPageToken'))
        listFiles(id, drive, folders, files, response.get('nextPageToken'))

def main():
    """Shows basic usage of the Docs API.
    Prints the title of a sample document.
    """

    creds = None
    # The file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.


    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)
#
#
    # Retrieve the documents contents from the Docs service.

    dr = build('drive', 'v3', credentials=creds)
    dr2 = build('drive', 'v2', credentials = creds)


    counter=0
    folders = []
    files = {}
    filesParents = {}
    if(os.path.exists('files.pickle')):
        with open('files.pickle', 'rb') as _files:
            files = pickle.load(_files)
    if(len(files)==0):    
        listFiles(DOCUMENT_ID, dr, folders, files, callingName = "",filesParents = filesParents)
        while(folders):
            newestFolder = folders.pop()
            print("Exploring folder %s"%newestFolder[1])
            listFiles(newestFolder[0], dr, folders, files, 
                callingName=newestFolder[1], filesParents = filesParents)

        with open('files.pickle', 'wb') as saved_file:
            pickle.dump(files, saved_file);
        #Get revision list
        counter = 0
        for a in files:
            counter+=1
            print("Processing %s\nid %s"%(filesParents[a], files[a]["id"]))
            print("%d out of %d"%(counter, len(files)))
            cur_id = files[a]["id"]
            revision = dr2.revisions().list(fileId = cur_id).execute().get("items", [])
            files[a]["revisions"]=revision

        with open('files.pickle', 'wb') as saved_file:
            print("safely dumped")
            pickle.dump(files, saved_file);

#Download revisions
    header = {}
    creds.apply(header)
    csvdata = {}
    minDate = datetime(9998, 12, 30, 23, 59, 59, 1000)
    maxDate = datetime(2, 12, 30, 23, 59, 59, 1000)
    counter =0
    for f in files:

        f=files[f]
        counter+=1
        print("%d out of %d"%(counter,len(files)))

        #TOBE FIXED
        fileName = f["id"]

        csvdata[fileName]={}


        dir_name =  DIR+slugify(f["name"])
        print(dir_name)
        p = Path(dir_name) 

        p.mkdir(exist_ok=True)
        for r in f["revisions"]:
            fname = dir_name+"/"+slugify(r["id"])+".txt"

            text=0
            if(not os.path.exists(fname)):
                response = requests.get(r["exportLinks"]["text/plain"], headers=header).content
                open(fname, 'wb').write(response)
                text=response.decode('utf-8-sig')
            else:
                text = open(fname, "rb").read()
                text = text.decode('utf-8-sig')
            text = str(text)
            flen = text.count(' ')
            moddate = iso8601.parse_date(r["modifiedDate"]).replace(tzinfo=None)
            csvdata[fileName][moddate]=flen

            minDate = min(minDate, moddate)
            maxDate = max(maxDate, moddate)


    csvdata_ = {}
    for a in csvdata:
        csvdata_[a] = OrderedDict()
        for key in sorted(csvdata[a]):
            csvdata_[a][key] = csvdata[a][key]
    csvdata_df = pd.DataFrame(csvdata)
    csvdata_df=csvdata_df.reset_index().sort_values('index').reset_index(drop=True)
    #pd.DataFrame(csvdata).T.to_excel('data.xlsx', sheet_name="csvdata")
    with open("csvdata.pickle", "wb") as cv:
        pickle.dump(csvdata, cv)





    '''
    ##Generate linspace Time
    inc = (maxDate-minDate)/BINS
    times = {}
    ##Add word count
    for f in csvdata:
        prevLen = 0
        for d in csvdata[f].keys():
            index_d = math.floor((d-minDate)/inc)*inc + minDate
            times.setdefault(index_d, {})
            times[index_d][f] = csvdata[f][d] - prevLen
           # times[index_d][f] = csvdata[f][d]
            prevLen = csvdata[f][d]

    stimes_ = OrderedDict()

    for key in sorted(times):
        stimes_[key] = times[key]
    times=stimes_

    df = pd.DataFrame(times)
    df.reindex(columns=df.columns[::-1])
    col = df.columns

    df.loc["sum of changes"]=0
    df.loc["total word count"] = 0
    for c in range(len(df.columns)):
        df.loc["sum of changes"].iloc[c] = df.iloc[:,c].sum()
        if(c>0):
            df.loc["total word count"].iloc[c] = df.loc["total word count"].iloc[c-1] \
            +df.loc["sum of changes"].iloc[c]
        else:
            df.loc["total word count"].iloc[c] = df.loc["sum of changes"].iloc[c]

    append_df_to_excel('text.xlsx', df, sheet_name="Change Data")
    pp.pprint(df)
    with open("dataframe.pickle", 'wb') as t:
        pickle.dump(df, t)
    print(maxDate, minDate, inc)

    with open("times.pickle", "wb") as t:
        pickle.dump(times, t)
    df.to_html('i.html')
'''

    append_df_to_excel('text.xlsx', csvdata_df, sheet_name = "1Total data")
    pp.pprint(csvdata_df)





    with open("csvdata.pickle", "wb") as cv:
        pickle.dump(csvdata, cv)

    




def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
if __name__ == '__main__':
    main()
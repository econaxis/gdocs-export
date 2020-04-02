
from __future__ import print_function
import pickle
import tkinter
import json
from slugify import slugify
import time
import math
from collections import OrderedDict
import os.path
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import pprint
from datetime import datetime, timezone, timedelta
import iso8601
import numpy
import matplotlib.pyplot as plt
import requests
import os.path
import pandas as pd
from pathlib import Path
pp = pprint.PrettyPrinter(indent=4);
BINS=400
readPermissions = False
# If modifying these scopes, delete the file token.pickle.
SCOPES = ['https://www.googleapis.com/auth/drive.metadata.readonly', 'https://www.googleapis.com/auth/drive.activity.readonly']
ITEMPERREQ = 1000

# The ID of a sample document.
DOCUMENT_ID = '0B4Fujvv5MfqbeTVRc3hIbXRfNE0'
MAX_FILES = 150
DIR = 'data/'


#Explores directory structure
def listRevisions (id, drive, folders, files, nextToken = None, callingName = "", filesParents = None):
    print("list %s"%id)
    q = "'" + id + "' in parents";
    if nextToken == None:
        response = drive.files().list(q=q, pageSize = 1000).execute()
    else:
        response = drive.files().list(q=q, pageSize = 1000, pageToken = nextToken).execute()
    for file in response.get('files', []):
        # Process change
        if(file.get("mimeType") == "application/vnd.google-apps.folder"):

            folders.append((file.get("id"),callingName+"/"+file.get("name")))
           # folders[file.get("id")] = files[file.get("name")] # Reference to scope of map to add
          
        elif (file.get("mimeType") == "application/vnd.google-apps.document"):
            files[file.get("id")] = file;
            if(filesParents!=None):
                filesParents[file.get("id")] = callingName + "/"+file.get("name")

    print("token %s"%response.get('nextPageToken'))
    if(response.get('nextPageToken')!=("" or None)):
        print("next token %s"%response.get('nextPageToken'))
        listRevisions(id, drive, folders, files, response.get('nextPageToken'))


def main():
    """Shows basic usage of the Docs API.
    Prints the title of a sample document.
    """

    creds = None
    # The file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.


    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)
#
#
    # Retrieve the documents contents from the Docs service.

    dr = build('drive', 'v3', credentials=creds)
    dr2 = build('drive', 'v2', credentials = creds)
    act = build('driveactivity', 'v2', credentials=creds)

    counter=0
    folders = []
    files = {}
    filesParents = {}
    if(os.path.exists('files.pickle')):
        with open('files.pickle', 'rb') as _files:
            files = pickle.load(_files)
    if(len(files)==0):    
        listRevisions(DOCUMENT_ID, dr, folders, files, callingName = "",filesParents = filesParents)
        while(folders):
            if(len(files)>MAX_FILES):
                break
            newestFolder = folders.pop()
            print("Exploring folder %s"%newestFolder[1])
            listRevisions(newestFolder[0], dr, folders, files, 
                callingName=newestFolder[1], filesParents = filesParents)

        #Get revision list
        counter = 0
        for a in files:
            counter+=1
            print("Processing %s"%(filesParents[a]))
            print("%d out of %d\n"%(counter, len(files)))
            cur_id = files[a]["id"]
            revision = dr2.revisions().list(fileId = cur_id).execute().get("items", [])
            files[a]["revisions"]=revision

        with open('files.pickle', 'wb') as saved_file:
            print("safely dumped")
            pickle.dump(files, saved_file);

#Download revisions

    #Apply creds to HTTP Header
    header = {}
    creds.apply(header)


    csvdata = {}
    counter = 0 

    #Should not run
    if (os.path.exists('csvdata.pickle') and False):
        with open('csvdata.pickle', 'rb') as cv:
            csvdata = pickle.load(cv)

    #Should run
    else:
        for f in files:
            f=files[f]

            API_WAIT=True
            while(API_WAIT):
                try:
                    drActivityResult = None
                    drActivityResult=act.activity().query(body={
                        'itemName' : "items/"+f["id"], 'pageSize' : 1000,
                        'filter' : "detail.action_detail_case: EDIT"
                     }).execute()
                    if(drActivityResult!=None):
                        API_WAIT=False
                except:
                    print("Google Docs API Limit Reached, Waiting...")
                    time.sleep(2)


            fActivities = []

            for a in drActivityResult.get("activities", []):
                fActivities.append(a.get("timestamp"))

            counter+=1
            print("%d out of %d"%(counter,len(files)))

            #TOBE FIXED
            fileName = f["name"]


            dir_name =  DIR+slugify(f["name"])
            print(dir_name)
            p = Path(dir_name) 

            p.mkdir(exist_ok=True)

            for r in fActivities:
                moddate = iso8601.parse_date(r).replace(tzinfo=None) - timedelta(hours = 7)
                csvdata[(fileName, moddate)] = 2


            for r in f["revisions"]:
                text=0
                if(readPermissions):
                    fname = dir_name+"/"+slugify(r["id"])+".txt"
                    if(not os.path.exists(fname)):
                        response = requests.get(r["exportLinks"]["text/plain"], headers=header).content
                        open(fname, 'wb').write(response)
                        text=response.decode('utf-8-sig')
                    else:
                        text = open(fname, "rb").read()
                        text = text.decode('utf-8-sig')
                    text= str(text)
                else: 
                    text = "Permission denied"
                flen = text.count(' ')
                moddate = iso8601.parse_date(r["modifiedDate"]).replace(tzinfo=None)

                #TIME ZONE CONVERSION
                moddate -= timedelta(hours = 7)
                csvdata[(fileName, moddate)] = flen
            #    csvdata[fileName][moddate]=flen


    pickle.dump(csvdata, open('unformatted_csvdata.pickle', 'wb'))
    MI = pd.MultiIndex.from_tuples(csvdata, names = ["Title", "Dates"])
    csvdata_df=pd.DataFrame(csvdata.values(),index = MI, columns = ["Type"])

    del csvdata, MI


    len_data = len(csvdata_df.index.levels[1])
    csvdata_df = pd.DataFrame(csvdata_df, columns = ["Type"])
    sumDatesIndex = pd.MultiIndex.from_product([["sumDates"], csvdata_df.index.levels[1]])
    sumDatesFrame = pd.DataFrame([1]*len_data, index = sumDatesIndex)

   

    csvdata_df = csvdata_df.append(sumDatesFrame)

    csvdata_df.to_pickle('csvdata.pickle')
    pickle.dump(csvdata_df, open('csvdata.pickle', 'wb'))

    activity_gen()



   
def activity_gen():
    csvdata = pd.read_pickle('csvdata.pickle')
    hists = {}
    activity = dict(time=[], files=[], marker_size=[])
    for f in csvdata.index.levels[0]:
        timesForFile = csvdata.loc[f].index
        activity["time"].append(csvdata.loc[f].index[-1])
        activity["files"].append(f)
        activity["marker_size"].append(log(len(timesForFile), 1.2))
        hists[f] = [0, 0]
        hists[f][0], bins = np.histogram([i.timestamp() for i in timesForFile], bins = 'auto')
        hists[f][1] = [datetime.fromtimestamp(i) for i in bins]
    pickle.dump(activity, open('activity.pickle', 'wb'))
    pickle.dump(hists, open('hists.pickle', 'wb'))

if __name__ == '__main__':
    main()

    '''

    #Sort dates per file
    print(1)
    csvdata_df = pd.DataFrame(csvdata)
    csvdata_df["Last Mod"] = 0
    for c in csvdata_df.columns:
        lastModDate=csvdata_df[c].dropna().index.tolist()[-1]
        csvdata_df.loc["Last Mod",c] = lastModDate
    print(1)

    csvdata_df.sort_values("Last Mod", axis = 1, inplace=True)
    print(1)

    #Sort dates    
    csvdata_df=csvdata_df.drop("Last Mod", axis = 0).sort_index().append(csvdata_df.loc["Last Mod"])
    print(1)
    #Sort files according to last modified



    csvdata_df.to_csv('csvdata.csv')
    csvdata_df.to_pickle('csvdata.pickle')




#    append_df_to_excel('text.xlsx', csvdata_df, 
 #       sheet_name = "1Total data", truncate_sheet=True)
  #  pp.pprint(csvdata_df)





#    with open("csvdata.pickle", "wb") as cv:
#        pickle.dump(csvdata, cv)

    




def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
          #  startrow = writer.book[sheet_name].max_row
          startrow = 0

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

    '''